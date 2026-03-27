"""
EAS Station - Emergency Alert System
Copyright (c) 2025-2026 Timothy Kramer (KR8MER)

This file is part of EAS Station.

EAS Station is dual-licensed software:
- GNU Affero General Public License v3 (AGPL-3.0) for open-source use
- Commercial License for proprietary use

You should have received a copy of both licenses with this software.
For more information, see LICENSE and LICENSE-COMMERCIAL files.

IMPORTANT: This software cannot be rebranded or have attribution removed.
See NOTICE file for complete terms.

Repository: https://github.com/KR8MER/eas-station
"""

from __future__ import annotations

"""Debugging endpoints for inspecting alerts and boundaries."""

from collections import OrderedDict

import io
from datetime import datetime

import pytz
from flask import Flask, jsonify, render_template, Response
from sqlalchemy import func

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app_core.extensions import db
from app_core.location import get_location_settings
from app_core.models import Boundary, CAPAlert, Intersection, PollDebugRecord
from app_core.poller_debug import ensure_poll_debug_table, serialise_debug_record, summarise_run
from app_utils.pdf_generator import generate_pdf_document

# Conversion factor: square metres → square miles
_SQM_PER_SQMI = 2_589_988.11


def register(app: Flask, logger) -> None:
    """Attach debug inspection routes to the Flask app."""

    route_logger = logger.getChild("routes_debug")

    def load_ipaws_debug_runs(max_runs: int = 10, record_limit: int = 600):
        """Return recent IPAWS poll debug runs for reuse across endpoints."""

        ensure_poll_debug_table(route_logger)
        location_settings = get_location_settings()
        timezone_name = (location_settings or {}).get("timezone") or "UTC"
        try:
            tz = pytz.timezone(timezone_name)
        except Exception:
            tz = pytz.UTC

        recent_records = (
            PollDebugRecord.query.order_by(
                PollDebugRecord.poll_started_at.desc(),
                PollDebugRecord.id.desc(),
            )
            .limit(record_limit)
            .all()
        )

        grouped: "OrderedDict[str, list[PollDebugRecord]]" = OrderedDict()
        for record in recent_records:
            run_records = grouped.get(record.poll_run_id)
            if run_records is None:
                if len(grouped) >= max_runs:
                    continue
                run_records = []
                grouped[record.poll_run_id] = run_records
            run_records.append(record)

        debug_runs = []
        for run_records in grouped.values():
            summary = summarise_run(run_records)
            poll_started_at = summary.get("poll_started_at")
            summary["poll_started_at_iso"] = (
                poll_started_at.isoformat() if poll_started_at else None
            )
            try:
                summary["poll_started_local"] = (
                    poll_started_at.astimezone(tz).strftime("%Y-%m-%d %H:%M:%S %Z")
                    if poll_started_at
                    else None
                )
            except Exception:
                summary["poll_started_local"] = summary["poll_started_at_iso"]
            summary["poll_started_display"] = (
                summary["poll_started_local"] or summary["poll_started_at_iso"]
            )

            alerts = []
            for record in run_records:
                alert_dict = serialise_debug_record(record)
                alert_dict["created_at_iso"] = (
                    record.created_at.isoformat() if record.created_at else None
                )
                try:
                    alert_dict["created_at_local"] = (
                        record.created_at.astimezone(tz).strftime("%Y-%m-%d %H:%M:%S %Z")
                        if record.created_at
                        else None
                    )
                except Exception:
                    alert_dict["created_at_local"] = alert_dict["created_at_iso"]
                alert_dict["created_at"] = alert_dict["created_at_iso"]

                sent_dt = alert_dict.get("alert_sent")
                if sent_dt:
                    alert_dict["alert_sent_iso"] = sent_dt.isoformat()
                    try:
                        alert_dict["alert_sent_local"] = sent_dt.astimezone(tz).strftime(
                            "%Y-%m-%d %H:%M:%S %Z"
                        )
                    except Exception:
                        alert_dict["alert_sent_local"] = alert_dict["alert_sent_iso"]
                else:
                    alert_dict["alert_sent_iso"] = None
                    alert_dict["alert_sent_local"] = None
                alert_dict["alert_sent"] = alert_dict["alert_sent_iso"]

                alert_dict["alert_sent_display"] = (
                    alert_dict["alert_sent_local"] or alert_dict["alert_sent_iso"]
                )
                alert_dict["created_at_display"] = (
                    alert_dict["created_at_local"] or alert_dict["created_at_iso"]
                )
                alerts.append(alert_dict)

            summary["alerts"] = alerts
            summary.setdefault("totals", {})["alerts"] = len(alerts)
            summary["totals"]["accepted"] = sum(
                1 for alert in alerts if alert.get("is_relevant")
            )
            summary["totals"]["saved"] = sum(
                1 for alert in alerts if alert.get("was_saved")
            )
            summary["totals"]["new_saved"] = sum(
                1 for alert in alerts if alert.get("was_new")
            )
            summary["totals"]["parse_failures"] = sum(
                1 for alert in alerts if not alert.get("parse_success")
            )
            debug_runs.append(summary)

        return debug_runs, location_settings, timezone_name


    @app.route("/debug/alert/<int:alert_id>")
    def debug_alert(alert_id: int):
        """Inspect intersections for a specific alert."""

        try:
            alert = CAPAlert.query.get_or_404(alert_id)

            geometry_info = {}
            try:
                geom_details = db.session.query(
                    func.ST_GeometryType(alert.geom).label("geom_type"),
                    func.ST_SRID(alert.geom).label("srid"),
                    func.ST_Area(alert.geom).label("area"),
                ).first()
                if geom_details:
                    geometry_info = {
                        "type": geom_details.geom_type,
                        "srid": geom_details.srid,
                        "area": float(geom_details.area) if geom_details.area else 0,
                    }
            except Exception as exc:  # pragma: no cover - defensive
                route_logger.error(
                    "Error retrieving geometry info for alert %s: %s", alert_id, exc
                )

            boundaries = Boundary.query.all()
            intersection_results = []

            for boundary in boundaries:
                try:
                    result = db.session.query(
                        func.ST_Intersects(alert.geom, boundary.geom).label(
                            "intersects"
                        ),
                        func.ST_Area(
                            func.ST_Intersection(alert.geom, boundary.geom).cast("geography")
                        ).label("area"),
                    ).first()

                    area_sqm = float(result.area) if result and result.area else 0
                    intersection_results.append(
                        {
                            "boundary_id": boundary.id,
                            "boundary_name": boundary.name,
                            "boundary_type": boundary.type,
                            "intersects": bool(result.intersects)
                            if result and result.intersects is not None
                            else False,
                            "intersection_area_sqm": area_sqm,
                            "intersection_area_sqmi": round(area_sqm / _SQM_PER_SQMI, 4),
                        }
                    )
                except Exception as exc:  # pragma: no cover - defensive
                    route_logger.error(
                        "Error checking intersection with boundary %s: %s",
                        boundary.id,
                        exc,
                    )

            existing_intersections = (
                db.session.query(Intersection)
                .filter_by(cap_alert_id=alert_id)
                .count()
            )
            boundaries_in_db = Boundary.query.count()

            debug_info = {
                "alert_id": alert_id,
                "alert_event": alert.event,
                "alert_area_desc": alert.area_desc,
                "has_geometry": alert.geom is not None,
                "geometry_info": geometry_info,
                "boundaries_in_db": boundaries_in_db,
                "existing_intersections": existing_intersections,
                "intersection_results": intersection_results,
                "intersections_found": len(
                    [result for result in intersection_results if result["intersects"]]
                ),
                "errors": [],
            }

            return jsonify(debug_info)
        except Exception as exc:  # pragma: no cover - defensive
            route_logger.error("Error debugging alert %s: %s", alert_id, exc)
            return jsonify({"error": f"Debug failed: {exc}"}), 500

    @app.route("/debug/ipaws")
    def debug_ipaws():
        """Render a diagnostics dashboard for the IPAWS poller output."""

        try:
            debug_runs, location_settings, timezone_name = load_ipaws_debug_runs()
            return render_template(
                "ipaws_debug.html",
                debug_runs=debug_runs,
                location_settings=location_settings,
                timezone_name=timezone_name,
            )
        except Exception as exc:  # pragma: no cover - defensive
            route_logger.error("Error rendering IPAWS debug page: %s", exc)
            return jsonify({"error": str(exc)}), 500

    @app.route("/debug/ipaws/export.xlsx")
    def debug_ipaws_export_excel():
        """Export recent IPAWS poll debug runs as an Excel workbook."""

        try:
            debug_runs, _, timezone_name = load_ipaws_debug_runs()
            if not debug_runs:
                return (
                    "<h1>No IPAWS debug data available</h1>"
                    "<p>No poll runs have been recorded yet.</p>"
                    "<p><a href='/debug/ipaws'>← Back to Debug</a></p>",
                    404,
                )

            workbook = Workbook()
            runs_sheet = workbook.active
            runs_sheet.title = "Poll Runs"
            runs_sheet.append(
                [
                    "Poll Run ID",
                    "Started",
                    "Status",
                    "Data Source",
                    "Alerts",
                    "Accepted",
                    "Saved",
                    "New Saved",
                    "Parse Failures",
                ]
            )

            for run in debug_runs:
                totals = run.get("totals", {})
                runs_sheet.append(
                    [
                        run.get("poll_run_id"),
                        run.get("poll_started_display"),
                        run.get("poll_status"),
                        run.get("data_source"),
                        totals.get("alerts", 0),
                        totals.get("accepted", 0),
                        totals.get("saved", 0),
                        totals.get("new_saved", 0),
                        totals.get("parse_failures", 0),
                    ]
                )

            alerts_sheet = workbook.create_sheet("Alerts")
            alerts_sheet.append(
                [
                    "Poll Run ID",
                    "Identifier",
                    "Event",
                    "Relevant",
                    "Saved",
                    "New",
                    "Parse Success",
                    "Parse Error",
                    "Sent",
                    "Captured",
                    "Relevance Reason",
                ]
            )

            for run in debug_runs:
                run_id = run.get("poll_run_id")
                for alert in run.get("alerts", []):
                    alerts_sheet.append(
                        [
                            run_id,
                            alert.get("alert_identifier"),
                            alert.get("alert_event"),
                            "Yes" if alert.get("is_relevant") else "No",
                            "Yes" if alert.get("was_saved") else "No",
                            "Yes" if alert.get("was_new") else "No",
                            "Yes" if alert.get("parse_success") else "No",
                            alert.get("parse_error") or "",
                            alert.get("alert_sent_display"),
                            alert.get("created_at_display"),
                            alert.get("relevance_reason") or "",
                        ]
                    )

            for sheet in (runs_sheet, alerts_sheet):
                for column_cells in sheet.columns:
                    max_length = max(
                        len(str(cell.value)) if cell.value is not None else 0
                        for cell in column_cells
                    )
                    adjusted_width = min(max_length + 2, 60)
                    sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)

            filename = f"ipaws_polls_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response = Response(
                output.getvalue(),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response.headers["Content-Disposition"] = f"attachment; filename={filename}"
            return response

        except Exception as exc:  # pragma: no cover - defensive
            route_logger.error("Error exporting IPAWS debug Excel: %s", exc)
            return (
                "<h1>Error exporting IPAWS debug data</h1>"
                f"<p>{exc}</p><p><a href='/debug/ipaws'>← Back to Debug</a></p>",
                500,
            )

    @app.route("/debug/ipaws/export.pdf")
    def debug_ipaws_export_pdf():
        """Export recent IPAWS poll debug runs as a PDF report."""

        try:
            debug_runs, _, timezone_name = load_ipaws_debug_runs()
            if not debug_runs:
                return (
                    "<h1>No IPAWS debug data available</h1>"
                    "<p>No poll runs have been recorded yet.</p>"
                    "<p><a href='/debug/ipaws'>← Back to Debug</a></p>",
                    404,
                )

            sections = []
            for run in debug_runs:
                totals = run.get("totals", {})
                heading = (
                    f"Poll {run.get('poll_run_id')}"
                    f" ({run.get('poll_started_display') or 'Unknown start'})"
                )
                content: list[str] = [
                    (
                        f"Status: {run.get('poll_status') or 'unknown'} | "
                        f"Data source: {run.get('data_source') or 'n/a'}"
                    ),
                    (
                        f"Alerts: {totals.get('alerts', 0)} | Accepted: {totals.get('accepted', 0)} | "
                        f"Saved: {totals.get('saved', 0)} | Parse failures: {totals.get('parse_failures', 0)}"
                    ),
                ]

                alerts = run.get("alerts", [])
                if alerts:
                    content.append("Alerts:")
                    for alert in alerts:
                        line = (
                            f"  - {alert.get('alert_identifier') or 'N/A'}"
                            f" ({alert.get('alert_event') or 'Unknown'}) | Relevant: "
                            f"{'yes' if alert.get('is_relevant') else 'no'} | Saved: "
                            f"{'yes' if alert.get('was_saved') else 'no'}"
                        )
                        if alert.get('parse_error'):
                            line += f" | Parse error: {alert.get('parse_error')}"
                        content.append(line)
                        if alert.get('relevance_reason'):
                            content.append(f"      Reason: {alert.get('relevance_reason')}")
                else:
                    content.append("No alerts captured for this run.")

                sections.append({'heading': heading, 'content': content})

            pdf_bytes = generate_pdf_document(
                title="IPAWS Poll Debug Export",
                sections=sections,
                subtitle=f"Timezone: {timezone_name}",
                footer_text="Generated by EAS Station - Emergency Alert System Platform",
            )

            filename = f"ipaws_polls_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
            response = Response(pdf_bytes, mimetype="application/pdf")
            response.headers["Content-Disposition"] = f"inline; filename={filename}"
            return response

        except Exception as exc:  # pragma: no cover - defensive
            route_logger.error("Error exporting IPAWS debug PDF: %s", exc)
            return (
                "<h1>Error exporting IPAWS debug data</h1>"
                f"<p>{exc}</p><p><a href='/debug/ipaws'>← Back to Debug</a></p>",
                500,
            )

    @app.route("/debug/boundaries/<int:alert_id>")
    def debug_boundaries(alert_id: int):
        """Debug boundary intersections for a specific alert."""

        try:
            alert = CAPAlert.query.get_or_404(alert_id)

            debug_info = {
                "alert_id": alert_id,
                "alert_event": alert.event,
                "alert_area_desc": alert.area_desc,
                "has_geometry": alert.geom is not None,
                "boundaries_in_db": Boundary.query.count(),
                "existing_intersections": Intersection.query.filter_by(
                    cap_alert_id=alert_id
                ).count(),
                "errors": [],
            }

            if not alert.geom:
                debug_info["errors"].append("Alert has no geometry data")
                return jsonify(debug_info)

            boundaries = Boundary.query.all()
            intersection_results = []

            for boundary in boundaries:
                try:
                    intersection_test = db.session.query(
                        func.ST_Intersects(alert.geom, boundary.geom).label(
                            "intersects"
                        ),
                        func.ST_Area(
                            func.ST_Intersection(alert.geom, boundary.geom).cast("geography")
                        ).label("area"),
                    ).first()

                    area_sqm = float(intersection_test.area) if intersection_test and intersection_test.area else 0
                    intersection_results.append(
                        {
                            "boundary_id": boundary.id,
                            "boundary_name": boundary.name,
                            "boundary_type": boundary.type,
                            "intersects": bool(intersection_test.intersects)
                            if intersection_test and intersection_test.intersects is not None
                            else False,
                            "intersection_area_sqm": area_sqm,
                            "intersection_area_sqmi": round(area_sqm / _SQM_PER_SQMI, 4),
                        }
                    )
                except Exception as exc:  # pragma: no cover - defensive
                    debug_info["errors"].append(
                        f"Error testing boundary {boundary.id}: {exc}"
                    )

            debug_info["intersection_results"] = intersection_results
            debug_info["intersections_found"] = len(
                [result for result in intersection_results if result["intersects"]]
            )

            try:
                geom_info = db.session.query(
                    func.ST_GeometryType(alert.geom).label("geom_type"),
                    func.ST_SRID(alert.geom).label("srid"),
                    func.ST_Area(alert.geom).label("area"),
                ).first()

                debug_info["geometry_info"] = {
                    "type": geom_info.geom_type if geom_info else "Unknown",
                    "srid": geom_info.srid if geom_info else "Unknown",
                    "area": float(geom_info.area)
                    if geom_info and geom_info.area
                    else 0,
                }
            except Exception as exc:  # pragma: no cover - defensive
                debug_info["errors"].append(
                    f"Error getting geometry info: {exc}"
                )

            return jsonify(debug_info)
        except Exception as exc:  # pragma: no cover - defensive
            route_logger.error("Error in debug_boundaries for %s: %s", alert_id, exc)
            return jsonify({"error": str(exc), "alert_id": alert_id}), 500


__all__ = ["register"]
